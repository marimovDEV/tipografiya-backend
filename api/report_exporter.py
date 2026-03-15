# backend/api/report_exporter.py
"""
Excel Report Exporter Service
Generates Excel files for various reports using openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q
from decimal import Decimal
import io

from .models import Order, ProductionStep, User, Material, MaterialBatch, WorkerTimeLog, QCCheckpoint


class ReportExporter:
    """Excel report generator with styling"""
    
    @staticmethod
    def _create_styled_workbook(title):
        """Create a workbook with standard styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel限制31个字符
        
        # Header style
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        return wb, ws, header_fill, header_font
    
    @staticmethod
    def _auto_adjust_columns(ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @classmethod
    def export_daily_production(cls, date_str):
        """
        Export daily production report
        Columns: Order, Client, Status, Progress, Deadline, Revenue
        """
        wb, ws, header_fill, header_font = cls._create_styled_workbook("Kunlik Ishlab Chiqarish")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Kunlik Ishlab Chiqarish Hisoboti - {date_str}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Buyurtma', 'Mijoz', 'Holat', 'Progress', 'Muddat', 'Narx']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_date=target_date,
            is_deleted=False
        ).select_related('client')
        
        row = 4
        total_revenue = 0
        
        for order in orders:
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=order.client.full_name if order.client else "-")
            ws.cell(row=row, column=3, value=order.get_status_display())
            
            # Progress
            total_steps = order.production_steps.count()
            completed = order.production_steps.filter(status='completed').count()
            progress = f"{completed}/{total_steps}" if total_steps > 0 else "0/0"
            ws.cell(row=row, column=4, value=progress)
            
            ws.cell(row=row, column=5, value=str(order.deadline) if order.deadline else "-")
            ws.cell(row=row, column=6, value=float(order.total_price or 0))
            
            total_revenue += float(order.total_price or 0)
            row += 1
        
        # Summary
        row += 1
        summary_cell = ws.cell(row=row, column=5)
        summary_cell.value = "JAMI:"
        summary_cell.font = Font(bold=True)
        
        total_cell = ws.cell(row=row, column=6)
        total_cell.value = total_revenue
        total_cell.font = Font(bold=True)
        
        cls._auto_adjust_columns(ws)
        
        return wb
    
    @classmethod
    def export_worker_efficiency(cls, start_date_str, end_date_str):
        """
        Export worker efficiency report
        Columns: Worker, Total Tasks, Completed, Work Hours, Efficiency %
        """
        wb, ws, header_fill, header_font = cls._create_styled_workbook("Xodimlar Samaradorligi")
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Xodimlar Samaradorligi - {start_date_str} to {end_date_str}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Xodim', 'Jami Vazifalar', 'Tugallangan', 'Ish Vaqti (soat)', 'Samaradorlik %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        users = User.objects.filter(is_active=True)
        row = 4
        
        for user in users:
            # Get steps for this user in date range
            steps = ProductionStep.objects.filter(
                assigned_to=user,
                created_at__gte=start_date,
                created_at__lte=end_date
            )
            
            total_steps = steps.count()
            if total_steps == 0:
                continue
            
            completed_steps = steps.filter(status='completed').count()
            
            # Calculate work hours from time logs
            time_logs = WorkerTimeLog.objects.filter(
                worker=user,
                timestamp__gte=start_date,
                timestamp__lte=end_date
            )
            
            total_minutes = 0
            current_start = None
            
            for log in time_logs.order_by('timestamp'):
                if log.action == 'start':
                    current_start = log.timestamp
                elif log.action == 'finish' and current_start:
                    duration = (log.timestamp - current_start).total_seconds() / 60
                    total_minutes += duration
                    current_start = None
            
            work_hours = total_minutes / 60
            efficiency = (completed_steps / total_steps * 100) if total_steps > 0 else 0
            
            ws.cell(row=row, column=1, value=user.username)
            ws.cell(row=row, column=2, value=total_steps)
            ws.cell(row=row, column=3, value=completed_steps)
            ws.cell(row=row, column=4, value=round(work_hours, 2))
            ws.cell(row=row, column=5, value=round(efficiency, 1))
            
            row += 1
        
        cls._auto_adjust_columns(ws)
        
        return wb
    
    @classmethod
    def export_warehouse_status(cls):
        """
        Export warehouse status report
        Columns: Material, Stock, Unit, Min Level, Status, Last Updated
        """
        wb, ws, header_fill, header_font = cls._create_styled_workbook("Sklad Holati")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Sklad Holati - {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Material', 'Zaxira', 'Birlik', 'Min. Daraja', 'Holat', 'Yangilangan']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        materials = Material.objects.filter(is_deleted=False)
        row = 4
        
        for material in materials:
            ws.cell(row=row, column=1, value=material.name)
            ws.cell(row=row, column=2, value=float(material.current_stock or 0))
            ws.cell(row=row, column=3, value=material.unit)
            ws.cell(row=row, column=4, value=float(material.minimum_stock or 0))
            
            # Status
            if material.current_stock < material.minimum_stock:
                status = "KAM QOLGAN"
                status_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            else:
                status = "OK"
                status_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            status_cell = ws.cell(row=row, column=5, value=status)
            status_cell.fill = status_fill
            
            ws.cell(row=row, column=6, value=material.updated_at.strftime('%Y-%m-%d'))
            
            row += 1
        
        cls._auto_adjust_columns(ws)
        
        return wb
    
    @classmethod
    def export_qc_statistics(cls, start_date_str, end_date_str):
        """
        Export QC statistics report
        Columns: Date, Total Checks, Passed, Failed, Pass Rate, Defects
        """
        wb, ws, header_fill, header_font = cls._create_styled_workbook("QC Statistikasi")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"QC Statistikasi - {start_date_str} to {end_date_str}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Sana', 'Jami', "O'tdi", "O'tmadi", 'Pass Rate %', 'Rework']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data by day
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        row = 4
        current_date = start_date
        
        while current_date <= end_date:
            checkpoints = QCCheckpoint.objects.filter(
                is_deleted=False,
                created_at__date=current_date.date()
            )
            
            total = checkpoints.count()
            if total == 0:
                current_date += timedelta(days=1)
                continue
            
            passed = checkpoints.filter(status='pass').count()
            failed = checkpoints.filter(status='fail').count()
            rework = checkpoints.filter(rework_triggered=True).count()
            pass_rate = (passed / total * 100) if total > 0 else 0
            
            ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=passed)
            ws.cell(row=row, column=4, value=failed)
            ws.cell(row=row, column=5, value=round(pass_rate, 1))
            ws.cell(row=row, column=6, value=rework)
            
            row += 1
            current_date += timedelta(days=1)
        
        cls._auto_adjust_columns(ws)
        
        return wb
    
    @classmethod
    def save_to_bytes(cls, workbook):
        """Save workbook to bytes for download"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
